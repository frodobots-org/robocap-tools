#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Calibration Validation Script
Download calibration results from S3, validate against acceptance criteria, and generate Excel report.
"""

import os
import sys
import argparse
import tempfile
import shutil
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
import glob

# Add script directory to path
script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

# Add s3sdk to path
project_root = Path(__file__).parent.parent
s3sdk_path = project_root / "s3sdk"
if str(s3sdk_path) not in sys.path:
    sys.path.insert(0, str(s3sdk_path))

try:
    from s3_sdk import S3SDK, S3Config
    from config_reader import load_s3_config
except ImportError as e:
    print(f"Error: Could not import S3 SDK: {e}")
    print("Please ensure s3sdk is available in the project root.")
    sys.exit(1)

# Import result parsers
from intrinsic_result_parser import get_intrinsic_calibration_params, find_intrinsic_result_file
from extrinsic_result_parser import get_extrinsic_imu_errors, find_extrinsic_result_file

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Error: openpyxl is required. Please install it with: pip install openpyxl")
    sys.exit(1)


@dataclass
class ValidationResult:
    """Validation result for a single calibration type"""
    device_id: str
    calibration_type: str  # 'intrinsic' or 'extrinsic'
    calibration_item: str  # 'front', 'eye', 'left', 'right'
    passed: bool
    reason: Optional[str] = None
    details: Optional[str] = None


class CalibrationValidator:
    """Validator for calibration results against acceptance criteria"""
    
    # Intrinsic validation ranges (from API_DOCUMENTATION.md)
    PROJECTION_RANGES = {
        'fx': (620.0, 660.0),
        'fy': (620.0, 660.0),
        'cx': (920.0, 1000.0),
        'cy': (500.0, 580.0)
    }
    
    DISTORTION_RANGES = {
        'k1': (0.01, 0.07),
        'k2': (-0.06, 0.06),
        'p1': (-0.06, 0.06),
        'p2': (-0.04, 0.04)
    }
    
    # Extrinsic validation thresholds
    REPROJECTION_ERROR_MAX = 1.5
    GYROSCOPE_ERROR_MAX = 0.015
    ACCELEROMETER_ERROR_MAX = 0.12
    
    def validate_intrinsic(
        self,
        device_id: str,
        calibration_item: str,
        params_dict: Dict
    ) -> Tuple[bool, Optional[str], Optional[str]]:
        """
        Validate intrinsic calibration parameters
        
        Args:
            device_id: Device ID
            calibration_item: Calibration item ('front', 'eye', 'left', 'right')
            params_dict: Parsed parameters dictionary (from intrinsic_result_parser)
            
        Returns:
            Tuple of (passed, reason, details)
        """
        if not params_dict:
            return False, "No parameters found", "Result file parsing returned empty"
        
        # Check for single camera (left/right) or dual camera (front/eye)
        is_dual = calibration_item in ['front', 'eye']
        
        cameras_to_check = ['cam0', 'cam1'] if is_dual else ['cam0']
        
        failures = []
        
        for cam_name in cameras_to_check:
            if cam_name not in params_dict:
                failures.append(f"Missing {cam_name} parameters")
                continue
            
            cam_data = params_dict[cam_name]
            
            # Validate projection parameters
            if 'projection' not in cam_data:
                failures.append(f"{cam_name}: Missing projection parameters")
                continue
            
            projection = cam_data['projection']
            if len(projection) != 4:
                failures.append(f"{cam_name}: Invalid projection length (expected 4, got {len(projection)})")
                continue
            
            fx, fy, cx, cy = projection
            
            # Check projection ranges
            if not (self.PROJECTION_RANGES['fx'][0] <= fx <= self.PROJECTION_RANGES['fx'][1]):
                failures.append(f"{cam_name}: fx={fx:.2f} out of range [{self.PROJECTION_RANGES['fx'][0]}, {self.PROJECTION_RANGES['fx'][1]}]")
            
            if not (self.PROJECTION_RANGES['fy'][0] <= fy <= self.PROJECTION_RANGES['fy'][1]):
                failures.append(f"{cam_name}: fy={fy:.2f} out of range [{self.PROJECTION_RANGES['fy'][0]}, {self.PROJECTION_RANGES['fy'][1]}]")
            
            if not (self.PROJECTION_RANGES['cx'][0] <= cx <= self.PROJECTION_RANGES['cx'][1]):
                failures.append(f"{cam_name}: cx={cx:.2f} out of range [{self.PROJECTION_RANGES['cx'][0]}, {self.PROJECTION_RANGES['cx'][1]}]")
            
            if not (self.PROJECTION_RANGES['cy'][0] <= cy <= self.PROJECTION_RANGES['cy'][1]):
                failures.append(f"{cam_name}: cy={cy:.2f} out of range [{self.PROJECTION_RANGES['cy'][0]}, {self.PROJECTION_RANGES['cy'][1]}]")
            
            # Validate distortion parameters
            if 'distortion' not in cam_data:
                failures.append(f"{cam_name}: Missing distortion parameters")
                continue
            
            distortion = cam_data['distortion']
            if len(distortion) != 4:
                failures.append(f"{cam_name}: Invalid distortion length (expected 4, got {len(distortion)})")
                continue
            
            k1, k2, p1, p2 = distortion
            
            # Check distortion ranges
            if not (self.DISTORTION_RANGES['k1'][0] <= k1 <= self.DISTORTION_RANGES['k1'][1]):
                failures.append(f"{cam_name}: k1={k1:.4f} out of range [{self.DISTORTION_RANGES['k1'][0]}, {self.DISTORTION_RANGES['k1'][1]}]")
            
            if not (self.DISTORTION_RANGES['k2'][0] <= k2 <= self.DISTORTION_RANGES['k2'][1]):
                failures.append(f"{cam_name}: k2={k2:.4f} out of range [{self.DISTORTION_RANGES['k2'][0]}, {self.DISTORTION_RANGES['k2'][1]}]")
            
            if not (self.DISTORTION_RANGES['p1'][0] <= p1 <= self.DISTORTION_RANGES['p1'][1]):
                failures.append(f"{cam_name}: p1={p1:.4f} out of range [{self.DISTORTION_RANGES['p1'][0]}, {self.DISTORTION_RANGES['p1'][1]}]")
            
            if not (self.DISTORTION_RANGES['p2'][0] <= p2 <= self.DISTORTION_RANGES['p2'][1]):
                failures.append(f"{cam_name}: p2={p2:.4f} out of range [{self.DISTORTION_RANGES['p2'][0]}, {self.DISTORTION_RANGES['p2'][1]}]")
        
        if failures:
            return False, "Parameters out of range", "; ".join(failures)
        
        return True, None, None
    
    def validate_extrinsic(
        self,
        device_id: str,
        calibration_item: str,
        errors_dict: Dict
    ) -> Tuple[bool, Optional[str], Optional[str]]:
        """
        Validate extrinsic calibration errors
        
        Args:
            device_id: Device ID
            calibration_item: Calibration item ('front', 'eye', 'left', 'right')
            errors_dict: Parsed errors dictionary (from extrinsic_result_parser)
            
        Returns:
            Tuple of (passed, reason, details)
        """
        if not errors_dict:
            return False, "No errors found", "Result file parsing returned empty"
        
        failures = []
        
        # Check reprojection errors
        is_dual = calibration_item in ['front', 'eye']
        
        # Check cam0 reprojection error
        if 'cam0_reprojection_error' not in errors_dict:
            failures.append("Missing cam0_reprojection_error")
        else:
            cam0_reproj = errors_dict['cam0_reprojection_error']
            if cam0_reproj >= self.REPROJECTION_ERROR_MAX:
                failures.append(f"cam0_reproj={cam0_reproj:.3f} >= {self.REPROJECTION_ERROR_MAX}")
        
        # Check cam1 reprojection error for dual camera
        if is_dual:
            if 'cam1_reprojection_error' not in errors_dict:
                failures.append("Missing cam1_reprojection_error")
            else:
                cam1_reproj = errors_dict['cam1_reprojection_error']
                if cam1_reproj >= self.REPROJECTION_ERROR_MAX:
                    failures.append(f"cam1_reproj={cam1_reproj:.3f} >= {self.REPROJECTION_ERROR_MAX}")
        
        # Check IMU errors (imu0, imu1, imu2)
        for imu_num in [0, 1, 2]:
            imu_key = f'imu{imu_num}'
            
            # Gyroscope errors
            gyro_mean_key = f'{imu_key}_gyroscope_error_mean'
            gyro_median_key = f'{imu_key}_gyroscope_error_median'
            
            if gyro_mean_key not in errors_dict:
                failures.append(f"Missing {gyro_mean_key}")
            else:
                gyro_mean = errors_dict[gyro_mean_key]
                if gyro_mean >= self.GYROSCOPE_ERROR_MAX:
                    failures.append(f"{gyro_mean_key}={gyro_mean:.5f} >= {self.GYROSCOPE_ERROR_MAX}")
            
            if gyro_median_key not in errors_dict:
                failures.append(f"Missing {gyro_median_key}")
            else:
                gyro_median = errors_dict[gyro_median_key]
                if gyro_median >= self.GYROSCOPE_ERROR_MAX:
                    failures.append(f"{gyro_median_key}={gyro_median:.5f} >= {self.GYROSCOPE_ERROR_MAX}")
            
            # Accelerometer errors
            accel_mean_key = f'{imu_key}_accelerometer_error_mean'
            accel_median_key = f'{imu_key}_accelerometer_error_median'
            
            if accel_mean_key not in errors_dict:
                failures.append(f"Missing {accel_mean_key}")
            else:
                accel_mean = errors_dict[accel_mean_key]
                if accel_mean >= self.ACCELEROMETER_ERROR_MAX:
                    failures.append(f"{accel_mean_key}={accel_mean:.4f} >= {self.ACCELEROMETER_ERROR_MAX}")
            
            if accel_median_key not in errors_dict:
                failures.append(f"Missing {accel_median_key}")
            else:
                accel_median = errors_dict[accel_median_key]
                if accel_median >= self.ACCELEROMETER_ERROR_MAX:
                    failures.append(f"{accel_median_key}={accel_median:.4f} >= {self.ACCELEROMETER_ERROR_MAX}")
        
        if failures:
            return False, "Errors exceed thresholds", "; ".join(failures)
        
        return True, None, None


class S3ResultsAnalyzer:
    """Analyzer for calibration results from S3"""
    
    def __init__(self, s3_sdk: S3SDK, temp_dir: str):
        """
        Initialize S3 results analyzer
        
        Args:
            s3_sdk: S3SDK instance
            temp_dir: Temporary directory for downloads
        """
        self.s3_sdk = s3_sdk
        self.temp_dir = temp_dir
        self.validator = CalibrationValidator()
        
        # Create temp directory if it doesn't exist
        os.makedirs(self.temp_dir, exist_ok=True)
    
    def list_all_devices(self) -> List[str]:
        """
        List all devices that have v1/results directory in S3
        
        Returns:
            List of device IDs
        """
        devices = []
        
        try:
            # List all files with v1/results prefix to find all devices
            # This is more reliable than listing folders
            all_files = self.s3_sdk.list_all_files(prefix='', recursive=True)
            
            # Extract unique device IDs from file paths
            device_set = set()
            
            for file_path in all_files:
                # Look for pattern: {device_id}/v1/results/
                if '/v1/results/' in file_path:
                    # Extract device ID (everything before /v1/results/)
                    parts = file_path.split('/v1/results/')
                    if parts:
                        device_id = parts[0].rstrip('/')
                        if device_id:
                            device_set.add(device_id)
            
            devices = sorted(list(device_set))
        
        except Exception as e:
            print(f"Error listing devices: {e}")
            # Fallback: try listing folders
            try:
                folders = self.s3_sdk.list_folders('')
                for folder in folders:
                    device_id = folder.rstrip('/')
                    results_path = f"{device_id}/v1/results/"
                    if self.s3_sdk.folder_exists(results_path):
                        if device_id not in devices:
                            devices.append(device_id)
                devices = sorted(devices)
            except Exception as e2:
                print(f"Fallback method also failed: {e2}")
        
        return devices
    
    def analyze_device(
        self,
        device_id: str
    ) -> Tuple[List[ValidationResult], List[ValidationResult]]:
        """
        Download and analyze a single device's calibration results
        
        Args:
            device_id: Device ID
            
        Returns:
            Tuple of (passed_results, failed_results)
        """
        passed_results = []
        failed_results = []
        
        # Create temporary directory for this device
        device_temp_dir = os.path.join(self.temp_dir, device_id)
        
        try:
            # Download v1/results directory
            s3_results_path = f"{device_id}/v1/results/"
            print(f"  Downloading {s3_results_path}...")
            
            if not self.s3_sdk.folder_exists(s3_results_path):
                print(f"  Warning: Results directory not found for {device_id}")
                failed_results.append(ValidationResult(
                    device_id=device_id,
                    calibration_type="unknown",
                    calibration_item="all",
                    passed=False,
                    reason="Results directory not found",
                    details="S3 path does not exist"
                ))
                return passed_results, failed_results
            
            download_results = self.s3_sdk.download_folder(s3_results_path, device_temp_dir)
            
            if not download_results:
                print(f"  Warning: No files downloaded for {device_id}")
                failed_results.append(ValidationResult(
                    device_id=device_id,
                    calibration_type="unknown",
                    calibration_item="all",
                    passed=False,
                    reason="Download failed",
                    details="No files found or download error"
                ))
                return passed_results, failed_results
            
            # Analyze intrinsic calibrations
            intrinsic_items = ['front', 'eye', 'left', 'right']
            for item in intrinsic_items:
                result = self._analyze_intrinsic(device_id, item, device_temp_dir)
                if result.passed:
                    passed_results.append(result)
                else:
                    failed_results.append(result)
            
            # Analyze extrinsic calibrations
            for item in intrinsic_items:
                result = self._analyze_extrinsic(device_id, item, device_temp_dir)
                if result.passed:
                    passed_results.append(result)
                else:
                    failed_results.append(result)
        
        except Exception as e:
            print(f"  Error analyzing device {device_id}: {e}")
            failed_results.append(ValidationResult(
                device_id=device_id,
                calibration_type="unknown",
                calibration_item="all",
                passed=False,
                reason="Analysis error",
                details=str(e)
            ))
        
        finally:
            # Clean up temporary directory
            if os.path.exists(device_temp_dir):
                try:
                    shutil.rmtree(device_temp_dir)
                    print(f"  Cleaned up temporary directory for {device_id}")
                except Exception as e:
                    print(f"  Warning: Failed to clean up {device_temp_dir}: {e}")
        
        return passed_results, failed_results
    
    def _analyze_intrinsic(
        self,
        device_id: str,
        item: str,
        results_dir: str
    ) -> ValidationResult:
        """Analyze intrinsic calibration for a specific item"""
        
        # Map item to task type for file finding
        task_type_map = {
            'front': 'cam_lr_front_intrinsic',
            'eye': 'cam_lr_eye_intrinsic',
            'left': 'cam_l_intrinsic',
            'right': 'cam_r_intrinsic'
        }
        
        task_type = task_type_map.get(item)
        if not task_type:
            return ValidationResult(
                device_id=device_id,
                calibration_type="intrinsic",
                calibration_item=item,
                passed=False,
                reason="Unknown item type",
                details=f"Invalid item: {item}"
            )
        
        # Try to find result file in results_dir and subdirectories
        result_file = None
        
        # First try in results_dir directly
        result_file = find_intrinsic_result_file(results_dir, task_type)
        
        # If not found, try in subdirectories
        if not result_file:
            subdirs = [
                'imus_cam_lr_front_extrinsic',
                'imus_cam_lr_eye_extrinsic',
                'imus_cam_l_extrinsic',
                'imus_cam_r_extrinsic'
            ]
            
            for subdir in subdirs:
                subdir_path = os.path.join(results_dir, subdir)
                if os.path.exists(subdir_path):
                    result_file = find_intrinsic_result_file(subdir_path, task_type)
                    if result_file:
                        break
        
        if not result_file:
            return ValidationResult(
                device_id=device_id,
                calibration_type="intrinsic",
                calibration_item=item,
                passed=False,
                reason="Result file not found",
                details=f"No intrinsic result file found for {item}"
            )
        
        # Parse parameters (use directory containing the result file)
        result_dir = os.path.dirname(result_file) if result_file else results_dir
        params_dict = get_intrinsic_calibration_params(result_dir, task_type)
        
        # Validate
        passed, reason, details = self.validator.validate_intrinsic(
            device_id, item, params_dict
        )
        
        return ValidationResult(
            device_id=device_id,
            calibration_type="intrinsic",
            calibration_item=item,
            passed=passed,
            reason=reason,
            details=details
        )
    
    def _analyze_extrinsic(
        self,
        device_id: str,
        item: str,
        results_dir: str
    ) -> ValidationResult:
        """Analyze extrinsic calibration for a specific item"""
        
        # Map item to task type for file finding
        task_type_map = {
            'front': 'cam_lr_front_extrinsic',
            'eye': 'cam_lr_eye_extrinsic',
            'left': 'cam_l_extrinsic',
            'right': 'cam_r_extrinsic'
        }
        
        task_type = task_type_map.get(item)
        if not task_type:
            return ValidationResult(
                device_id=device_id,
                calibration_type="extrinsic",
                calibration_item=item,
                passed=False,
                reason="Unknown item type",
                details=f"Invalid item: {item}"
            )
        
        # Try to find result file in results_dir and subdirectories
        result_file = None
        
        # First try in results_dir directly
        result_file = find_extrinsic_result_file(results_dir, task_type)
        
        # If not found, try in subdirectories
        if not result_file:
            subdirs = [
                'imus_cam_lr_front_extrinsic',
                'imus_cam_lr_eye_extrinsic',
                'imus_cam_l_extrinsic',
                'imus_cam_r_extrinsic'
            ]
            
            for subdir in subdirs:
                subdir_path = os.path.join(results_dir, subdir)
                if os.path.exists(subdir_path):
                    result_file = find_extrinsic_result_file(subdir_path, task_type)
                    if result_file:
                        break
        
        if not result_file:
            return ValidationResult(
                device_id=device_id,
                calibration_type="extrinsic",
                calibration_item=item,
                passed=False,
                reason="Result file not found",
                details=f"No extrinsic result file found for {item}"
            )
        
        # Parse errors (use directory containing the result file)
        result_dir = os.path.dirname(result_file) if result_file else results_dir
        errors_dict = get_extrinsic_imu_errors(result_dir, task_type)
        
        # Validate
        passed, reason, details = self.validator.validate_extrinsic(
            device_id, item, errors_dict
        )
        
        return ValidationResult(
            device_id=device_id,
            calibration_type="extrinsic",
            calibration_item=item,
            passed=passed,
            reason=reason,
            details=details
        )


class ExcelReporter:
    """Excel report generator"""
    
    def __init__(self, output_file: str):
        """
        Initialize Excel reporter
        
        Args:
            output_file: Output Excel file path
        """
        self.output_file = output_file
        self.workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Create sheets
        self.passed_sheet = self.workbook.create_sheet("Passed")
        self.failed_sheet = self.workbook.create_sheet("Failed")
        
        # Setup headers
        self._setup_headers()
    
    def _setup_headers(self):
        """Setup Excel sheet headers"""
        # Passed sheet headers
        passed_headers = ["Device ID", "Intrinsic Types", "Extrinsic Types", "All Passed"]
        self.passed_sheet.append(passed_headers)
        self._format_header_row(self.passed_sheet[1])
        
        # Failed sheet headers
        failed_headers = ["Device ID", "Failed Type", "Failed Item", "Reason", "Details"]
        self.failed_sheet.append(failed_headers)
        self._format_header_row(self.failed_sheet[1])
    
    def _format_header_row(self, row):
        """Format header row with bold font and background color"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
    
    def add_results(self, passed_results: List[ValidationResult], failed_results: List[ValidationResult]):
        """
        Add validation results to Excel report
        
        Args:
            passed_results: List of passed validation results
            failed_results: List of failed validation results
        """
        # Group passed results by device
        passed_by_device = {}
        for result in passed_results:
            if result.device_id not in passed_by_device:
                passed_by_device[result.device_id] = {
                    'intrinsic': [],
                    'extrinsic': []
                }
            
            if result.calibration_type == 'intrinsic':
                passed_by_device[result.device_id]['intrinsic'].append(result.calibration_item)
            else:
                passed_by_device[result.device_id]['extrinsic'].append(result.calibration_item)
        
        # Add passed devices to sheet
        for device_id, types in passed_by_device.items():
            intrinsic_types = ", ".join(sorted(types['intrinsic'])) if types['intrinsic'] else "None"
            extrinsic_types = ", ".join(sorted(types['extrinsic'])) if types['extrinsic'] else "None"
            self.passed_sheet.append([
                device_id,
                intrinsic_types,
                extrinsic_types,
                "Yes"
            ])
        
        # Add failed results to sheet
        for result in failed_results:
            self.failed_sheet.append([
                result.device_id,
                result.calibration_type,
                result.calibration_item,
                result.reason or "Unknown",
                result.details or ""
            ])
    
    def save(self):
        """Save Excel file"""
        self.workbook.save(self.output_file)
        print(f"\nExcel report saved to: {self.output_file}")


def main():
    """Main function"""
    parser = argparse.ArgumentParser(
        description='Validate calibration results from S3 and generate Excel report'
    )
    
    parser.add_argument(
        '--config',
        type=str,
        default='s3_config.json',
        help='Path to S3 configuration file (default: s3_config.json)'
    )
    
    parser.add_argument(
        '--output',
        type=str,
        default='calibration_validation_report.xlsx',
        help='Output Excel file path (default: calibration_validation_report.xlsx)'
    )
    
    parser.add_argument(
        '--temp-dir',
        type=str,
        default=None,
        help='Temporary directory for downloads (default: system temp directory)'
    )
    
    parser.add_argument(
        '--device-id',
        type=str,
        default=None,
        help='Single device ID to validate (optional, if not provided, validates all devices)'
    )
    
    args = parser.parse_args()
    
    # Setup temp directory
    if args.temp_dir:
        temp_dir = args.temp_dir
        os.makedirs(temp_dir, exist_ok=True)
    else:
        temp_dir = tempfile.mkdtemp(prefix='calibration_validation_')
    
    print(f"Temporary directory: {temp_dir}")
    
    # Load S3 config
    try:
        config = load_s3_config(args.config)
        s3_sdk = S3SDK(config)
        print(f"S3 SDK initialized with bucket: {config.bucket_name}")
    except Exception as e:
        print(f"Error loading S3 config: {e}")
        return 1
    
    # Initialize analyzer
    analyzer = S3ResultsAnalyzer(s3_sdk, temp_dir)
    
    # Get device list
    if args.device_id:
        devices = [args.device_id]
        print(f"Validating single device: {args.device_id}")
    else:
        print("Listing all devices from S3...")
        devices = analyzer.list_all_devices()
        print(f"Found {len(devices)} devices")
    
    if not devices:
        print("No devices found to validate")
        return 1
    
    # Initialize Excel reporter
    reporter = ExcelReporter(args.output)
    
    # Analyze all devices
    all_passed_results = []
    all_failed_results = []
    
    print(f"\n{'='*80}")
    print(f"Starting validation for {len(devices)} device(s)")
    print(f"{'='*80}\n")
    
    for i, device_id in enumerate(devices, 1):
        print(f"[{i}/{len(devices)}] Analyzing device: {device_id}")
        
        passed, failed = analyzer.analyze_device(device_id)
        
        all_passed_results.extend(passed)
        all_failed_results.extend(failed)
        
        passed_count = len(passed)
        failed_count = len(failed)
        print(f"  Results: {passed_count} passed, {failed_count} failed")
    
    # Generate report
    print(f"\n{'='*80}")
    print("Generating Excel report...")
    print(f"{'='*80}")
    
    reporter.add_results(all_passed_results, all_failed_results)
    reporter.save()
    
    # Summary
    print(f"\n{'='*80}")
    print("Validation Summary")
    print(f"{'='*80}")
    print(f"Total devices analyzed: {len(devices)}")
    print(f"Total passed validations: {len(all_passed_results)}")
    print(f"Total failed validations: {len(all_failed_results)}")
    print(f"Report saved to: {args.output}")
    print(f"{'='*80}\n")
    
    # Cleanup temp directory if we created it
    if not args.temp_dir and os.path.exists(temp_dir):
        try:
            shutil.rmtree(temp_dir)
            print(f"Cleaned up temporary directory: {temp_dir}")
        except Exception as e:
            print(f"Warning: Failed to clean up temp directory: {e}")
    
    return 0


if __name__ == '__main__':
    sys.exit(main())
